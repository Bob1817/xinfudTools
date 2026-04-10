import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime


class ReportGenerator:
    """个税申报表 Excel 生成器"""

    def generate(self, data: pd.DataFrame, period: str, output_dir: str) -> str:
        """生成申报表 Excel，返回文件路径"""
        Path(output_dir).mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"个税申报表_{period}_{timestamp}.xlsx"
        filepath = str(Path(output_dir) / filename)

        try:
            with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
                data.to_excel(writer, index=False, sheet_name="扣缴申报表")
                self._format_sheet(writer.book["扣缴申报表"])
        except Exception as e:
            raise RuntimeError(f"生成 Excel 文件失败：{e}")

        return filepath

    def _format_sheet(self, ws):
        """美化格式"""

        # 样式定义
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        data_font = Font(size=10)
        data_alignment = Alignment(vertical="center")
        center_alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 表头行
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 表头行高
        ws.row_dimensions[1].height = 35

        # 数据行
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            for cell in row:
                cell.font = data_font
                cell.border = thin_border

                # 工号、姓名、证件类型居中
                if cell.column <= 4:
                    cell.alignment = center_alignment
                else:
                    cell.alignment = data_alignment

                # 数字列格式化
                if isinstance(cell.value, (int, float)) and cell.column >= 5:
                    cell.number_format = '#,##0.00'

            # 交替行背景色
            if row_idx % 2 == 0:
                light_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
                for cell in row:
                    cell.fill = light_fill

        # 自动列宽
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0

            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20)):
                cell = row[col_idx - 1]
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            # 设置列宽，最小10，最大30
            ws.column_dimensions[col_letter].width = min(max(max_length + 4, 10), 30)

        # 冻结首行
        ws.freeze_panes = "A2"
